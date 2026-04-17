from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.base_generator import BaseReportGenerator
from reports.models import ReportData

_HEADER_COLUMNS = [
    "modelo", "marca", "status task", "detalle",
    "scheduled_at", "started_at", "end_at",
]

_DATE_FMT = "%Y-%m-%d %H:%M"
_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_ALT_FILL = PatternFill("solid", fgColor="ECF0F1")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt(value: datetime | None) -> str:
    return value.strftime(_DATE_FMT) if value else "-"


class ExcelReportGenerator(BaseReportGenerator):

    def generate(self, data: ReportData) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # ── Metadata block ──
        meta_rows = [
            ("MASSIVE TASK PROGRAM",),
            ("nombre tarea:", data.task_name),
            ("total devices:", data.total_devices),
            ("processed devices:", data.processed_devices),
            ("failed devices:", data.failed_devices),
            ("group name:", data.group_name),
            (),  # blank separator
        ]
        for row_data in meta_rows:
            ws.append(list(row_data))

        title_cell = ws["A1"]
        title_cell.font = Font(bold=True, size=14)

        for row_idx in range(2, 7):
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

        # ── Table header ──
        header_row_idx = ws.max_row + 1
        header = [""] + _HEADER_COLUMNS
        ws.append(header)
        for col_idx, _ in enumerate(header, start=1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # ── Data rows ──
        for i, row in enumerate(data.rows):
            ws.append([
                row.serial_number,
                row.model or "-",
                row.manufacturer or "-",
                row.task_status,
                row.detail or "-",
                _fmt(row.scheduled_at),
                _fmt(row.started_at),
                _fmt(row.end_at),
            ])
            fill = _ALT_FILL if i % 2 else _WHITE_FILL
            data_row_idx = header_row_idx + 1 + i
            for col_idx in range(1, len(header) + 1):
                ws.cell(row=data_row_idx, column=col_idx).fill = fill

        # ── Auto column width ──
        detail_col = 5
        for col_idx in range(1, len(header) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(header_row_idx, ws.max_row + 1)
            )
            max_width = 60 if col_idx == detail_col else 40
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, max_width)

        for r in range(header_row_idx + 1, ws.max_row + 1):
            ws.cell(row=r, column=detail_col).alignment = Alignment(wrap_text=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
